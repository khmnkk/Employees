import csv
from openpyxl import Workbook
from datetime import datetime

# Функція для обчислення віку на поточну дату
def calculate_age(birthdate):
    today = datetime.now().date()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

# Функція для створення XLSX файлу
def create_xlsx_from_csv(csv_filename, xlsx_filename):
    try:
        # Читаємо CSV файл
        try:
            with open(csv_filename, mode='r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                records = [row for row in reader]
        except FileNotFoundError:
            print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
            return

        # Перетворюємо дати народження та обчислюємо вік
        for row in records:
            try:
                birthdate = datetime.strptime(row['Дата народження'], '%Y-%m-%d').date()
                row['Вік'] = calculate_age(birthdate)
            except ValueError:
                row['Вік'] = None  # Якщо дата некоректна, віку не буде

        # Створюємо новий Excel файл
        try:
            wb = Workbook()

            # Створюємо аркуш "all"
            ws_all = wb.active
            ws_all.title = "all"
            ws_all.append(list(records[0].keys()))  # Заголовки колонок
            for row in records:
                ws_all.append(list(row.values()))

            # Фільтруємо записи за віковими категоріями
            def filter_records(min_age, max_age=None):
                return [r for r in records if r['Вік'] is not None and (r['Вік'] >= min_age) and (r['Вік'] <= max_age if max_age else r['Вік'] >= min_age)]

            # Створюємо аркуш для кожної вікової категорії
            def create_sheet(sheet_name, min_age, max_age=None):
                ws = wb.create_sheet(title=sheet_name)
                ws.append(list(records[0].keys()))  # Заголовки колонок
                filtered_records = filter_records(min_age, max_age)
                for row in filtered_records:
                    ws.append(list(row.values()))

            create_sheet("younger_18", 0, 17)
            create_sheet("18-45", 18, 45)
            create_sheet("45-70", 46, 70)
            create_sheet("older_70", 71)

            wb.save(xlsx_filename)
            print("Ok, програма завершила свою роботу успішно.")
        except Exception as e:
            print(f"Повідомлення про неможливість створення XLSX файлу: {e}")

    except Exception as e:
        print(f"Виникла помилка: {e}")

# Назва CSV файлу (створеного раніше)
csv_filename = 'people.csv'

# Назва XLSX файлу, який потрібно створити
xlsx_filename = 'people_by_age.xlsx'

# Викликаємо функцію для створення XLSX
create_xlsx_from_csv(csv_filename, xlsx_filename)
